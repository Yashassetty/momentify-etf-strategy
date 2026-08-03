import os
import sys
import re
import io
import json
import csv
import webbrowser
import datetime
import pandas as pd
import yfinance as yf

# Column Mapping dictionaries for normalization
TRADES_MAP = {
    'Stock Name': ['stock name', 'stock', 'ticker', 'symbol', 'etf'],
    'Entry Price': ['entry price', 'entry', 'entry_price', 'buy price', 'buy_price'],
    'Exit Price': ['exit price', 'exit', 'exit_price', 'sell price', 'sell_price'],
    'Qty': ['qty', 'quantity', 'count', 'shares'],
    'P&L': ['p&l', 'pnl', 'profit', 'gain', 'loss', 'realised pnl', 'realised p&l'],
    'Entry Date': ['entry date', 'entry_date', 'buy date', 'buy_date'],
    'Exit Date': ['exit date', 'exit_date', 'sell date', 'sell_date'],
    'Exit Reason': ['exit reason', 'exit_reason', 'reason']
}

TIMELINE_MAP = {
    'Date': ['date', 'time', 'week', 'day', 'timeline date'],
    'Unrealised PL': ['unrealised pl', 'unrealised pnl', 'unrealised', 'unrealized pl', 'unrealized pnl', 'unrealized'],
    'Realised PL': ['realised pl', 'realised pnl', 'realised', 'realized pl', 'realized pnl', 'realized'],
    'Total PL': ['total pl', 'total pnl', 'total', 'total_pl', 'total_pnl']
}

def clean_date_str(d_str):
    if not isinstance(d_str, str):
        return d_str
    # remove anything in parentheses (e.g. "(End of Period)")
    d_str = re.sub(r'\(.*?\)', '', d_str)
    return d_str.strip()

def normalize_dataframe_headers(df, expected_map):
    new_cols = {}
    for col in df.columns:
        col_clean = str(col).strip().lower().replace('_', ' ').replace('-', ' ')
        matched = False
        for std_name, alt_names in expected_map.items():
            if col_clean in alt_names:
                new_cols[col] = std_name
                matched = True
                break
        if not matched:
            new_cols[col] = str(col).strip()
    return df.rename(columns=new_cols)

def load_file_to_rows(file_path):
    rows = []
    if file_path.lower().endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Default to first sheet, or search for sheet names containing Trades/Timeline
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        for r in sheet.iter_rows(values_only=True):
            row_vals = [str(x).strip() if x is not None else "" for x in r]
            if all(v == "" for v in row_vals):
                rows.append([])
            else:
                rows.append(row_vals)
    else:
        # Read CSV with fallback encodings
        content = ""
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'utf-16']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except Exception:
                continue
        if not content:
            raise ValueError(f"Could not read the CSV file {file_path} with standard encodings.")
        
        reader = csv.reader(io.StringIO(content))
        for row in reader:
            rows.append(row)
    return rows

def parse_stacked_rows(rows):
    trades_header_idx = -1
    timeline_header_idx = -1
    
    # Identify table header indices
    for i, row in enumerate(rows):
        if not row:
            continue
        row_str = [str(x).strip().lower() for x in row if x is not None]
        if any("stock name" in s or s == "stock" for s in row_str) and any("entry price" in s or s == "entry" for s in row_str):
            trades_header_idx = i
        elif any("unrealised pl" in s or s == "unrealised" for s in row_str) and any("realised pl" in s or s == "realised" for s in row_str):
            timeline_header_idx = i
            
    if trades_header_idx == -1:
        raise ValueError("Could not find Trades Log table in the file (missing 'Stock Name' and 'Entry Price' headers).")
        
    # Extract Trades
    trades_data = []
    for i in range(trades_header_idx + 1, len(rows)):
        row = rows[i]
        if not row or all(str(x).strip() == "" for x in row):
            break
        if i == timeline_header_idx:
            break
        trades_data.append(row)
        
    # Extract Timeline
    timeline_data = []
    if timeline_header_idx != -1:
        for i in range(timeline_header_idx + 1, len(rows)):
            row = rows[i]
            row_str = "".join([str(x) for x in row if x is not None]).strip().lower()
            if "total p&l perfomance" in row_str or "realised p&l perfomance" in row_str:
                break
            if not row or all(str(x).strip() == "" for x in row):
                # Check next row to see if we reached performance section or end
                if i + 1 < len(rows):
                    next_str = "".join([str(x) for x in rows[i+1] if x is not None]).strip().lower()
                    if "total p&l" in next_str or "performance" in next_str or next_str == "":
                        break
                break
            timeline_data.append(row)
            
    # Load into DataFrames
    df_trades = pd.DataFrame(trades_data)
    if not df_trades.empty:
        df_trades.columns = [str(x).strip() for x in rows[trades_header_idx][:df_trades.shape[1]]]
        
    df_timeline = pd.DataFrame(timeline_data)
    if not df_timeline.empty and timeline_header_idx != -1:
        df_timeline.columns = [str(x).strip() for x in rows[timeline_header_idx][:df_timeline.shape[1]]]
        
    return df_trades, df_timeline

def parse_excel_sheets(file_path):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    trades_sheet = None
    timeline_sheet = None
    
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "trade" in name_lower or "log" in name_lower:
            trades_sheet = name
        elif "timeline" in name_lower or "trace" in name_lower or "pnl" in name_lower or "performance" in name_lower:
            timeline_sheet = name
            
    if trades_sheet and timeline_sheet:
        print(f"[INFO] Found multi-sheet Excel setup. Trades='{trades_sheet}', Timeline='{timeline_sheet}'")
        df_trades = sheet_to_df(wb[trades_sheet])
        df_timeline = sheet_to_df(wb[timeline_sheet])
        return df_trades, df_timeline
    return None

def sheet_to_df(sheet):
    rows = []
    for r in sheet.iter_rows(values_only=True):
        row_vals = [str(x).strip() if x is not None else "" for x in r]
        if any(v != "" for v in row_vals):
            rows.append(row_vals)
    if not rows:
        return pd.DataFrame()
    
    # Header is the first non-empty row
    header = rows[0]
    data = rows[1:]
    return pd.DataFrame(data, columns=header)

def parse_input_file(file_path):
    df_trades = None
    df_timeline = None
    
    if file_path.lower().endswith(('.xlsx', '.xls')):
        try:
            res = parse_excel_sheets(file_path)
            if res is not None:
                df_trades, df_timeline = res
        except Exception as e:
            print(f"[WARNING] Sheet-based Excel parsing failed: {e}. Falling back to stacked table parsing.")
            
    if df_trades is None or df_timeline is None:
        print("[INFO] Parsing stacked tables from file...")
        rows = load_file_to_rows(file_path)
        df_trades, df_timeline = parse_stacked_rows(rows)
        
    # Standardize column headers
    df_trades = normalize_dataframe_headers(df_trades, TRADES_MAP)
    if not df_timeline.empty:
        df_timeline = normalize_dataframe_headers(df_timeline, TIMELINE_MAP)
        
    return df_trades, df_timeline

def main():
    # 1. Identify input file
    input_file = None
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        # Auto-detect the first csv or xlsx file in the current directory (excluding output/scratch files)
        candidates = []
        for file in os.listdir('.'):
            if file.lower().endswith(('.csv', '.xlsx')) and file not in ['etf_closing_prices.csv', 'total_pnl_race.csv', 'realized_pnl_race.csv', 'mtm_pnl_race.csv']:
                candidates.append(file)
        if candidates:
            # Prefer xlsx if both exist
            candidates.sort(key=lambda x: 1 if x.endswith('.xlsx') else 2)
            input_file = candidates[0]
            print(f"[INFO] Auto-detected input file: {input_file}")
        else:
            print("[ERROR] No input CSV or Excel file provided as argument or found in directory.")
            print("Usage: python process_portfolio.py <path_to_excel_or_csv>")
            sys.exit(1)
            
    # 2. Parse the Excel / CSV file
    print(f"[INFO] Loading and parsing {input_file}...")
    df_trades, df_timeline = parse_input_file(input_file)
    
    # 3. Clean and convert data types
    print("[INFO] Cleaning trade data columns...")
    df_trades['Entry Date'] = pd.to_datetime(df_trades['Entry Date'].apply(clean_date_str), format='%d %B %Y')
    df_trades['Exit Date'] = pd.to_datetime(df_trades['Exit Date'].apply(clean_date_str), format='%d %B %Y')
    
    df_trades['Qty'] = pd.to_numeric(df_trades['Qty'])
    df_trades['Entry Price'] = pd.to_numeric(df_trades['Entry Price'])
    df_trades['Exit Price'] = pd.to_numeric(df_trades['Exit Price'])
    df_trades['P&L'] = pd.to_numeric(df_trades['P&L'])
    
    # Parse timeline dates
    if df_timeline.empty or 'Date' not in df_timeline.columns:
        print("[WARNING] Timeline table not found in input. Constructing weekly timeline from trade dates...")
        min_date = df_trades['Entry Date'].min()
        max_date = df_trades['Exit Date'].max()
        if pd.isna(max_date):
            max_date = pd.Timestamp.today()
        timeline_dates_parsed = pd.date_range(start=min_date, end=max_date, freq='W-MON')
    else:
        df_timeline['Date'] = pd.to_datetime(df_timeline['Date'].apply(clean_date_str), format='%d %B %Y')
        timeline_dates_parsed = df_timeline['Date'].dropna()
        
    # Merge timeline dates with trade entry/exit dates to get a 100% complete timeline
    trade_dates = set(df_trades['Entry Date'].dropna()).union(set(df_trades['Exit Date'].dropna()))
    merged_dates = set(timeline_dates_parsed).union(trade_dates)
    
    today = pd.Timestamp.today()
    valid_dates = sorted([d for d in merged_dates if d <= today])
    valid_date_strings = [d.strftime("%Y-%m-%d") for d in valid_dates]
    
    print(f"[INFO] Total unique timeline dates (including trade checkpoints): {len(valid_date_strings)}")
    
    # 4. Get unique ETFs and build Yahoo Finance tickers list
    unique_etfs = sorted(list(df_trades['Stock Name'].unique()))
    tickers = [f"{etf}.NS" for etf in unique_etfs]
    
    # Always include NIFTYBEES.NS for benchmark comparison
    benchmark_ticker = "NIFTYBEES.NS"
    download_tickers = list(set(tickers + [benchmark_ticker]))
    
    # 5. Fetch historical prices from Yahoo Finance
    start_fetch = (valid_dates[0] - pd.Timedelta(days=7)).strftime("%Y-%m-%d")
    end_fetch = (today + pd.Timedelta(days=2)).strftime("%Y-%m-%d")
    
    print(f"[INFO] Fetching historical closing prices from Yahoo Finance for {len(download_tickers)} tickers...")
    print(f"[INFO] Fetching range: {start_fetch} to {end_fetch}")
    
    prices_data = yf.download(download_tickers, start=start_fetch, end=end_fetch, progress=False)
    
    if isinstance(prices_data.columns, pd.MultiIndex):
        close_df = prices_data.xs('Close', level='Price', axis=1)
    else:
        close_df = prices_data['Close']
        
    # Reindex over valid timeline dates, forward and backward filling price gaps
    close_df = close_df.reindex(valid_date_strings).ffill().bfill()
    close_df.index = pd.to_datetime(close_df.index)
    
    # Save closing prices to file
    close_df_transposed = close_df.T
    close_df_transposed.index.name = "ETF"
    close_df_transposed.columns = pd.to_datetime(close_df_transposed.columns).strftime("%d-%b-%Y")
    close_df_transposed.to_csv("etf_closing_prices.csv")
    print("[INFO] Prices saved to etf_closing_prices.csv")
    
    # 6. Calculate realised and unrealised (MtM) P&L over timeline
    print("[INFO] Computing cumulative P&L trajectory for each stock...")
    total_pnl_df = pd.DataFrame(0.0, index=close_df.index, columns=unique_etfs)
    
    for date in close_df.index:
        for stock in unique_etfs:
            ticker = f"{stock}.NS"
            stock_trades = df_trades[df_trades['Stock Name'] == stock]
            total_pnl_sum = 0.0
            
            for _, trade in stock_trades.iterrows():
                entry_date = trade['Entry Date']
                exit_date = trade['Exit Date']
                qty = trade['Qty']
                entry_price = trade['Entry Price']
                final_pnl = trade['P&L']
                
                # Case 1: Realized P&L on or after Exit Date
                if date >= exit_date:
                    total_pnl_sum += final_pnl
                # Case 2: Mark-to-market (MtM) Unrealized P&L between Entry and Exit Date
                elif date >= entry_date:
                    current_price = None
                    if ticker in close_df.columns:
                        current_price = close_df.loc[date, ticker]
                    
                    # Fallback if Yahoo Finance price is unavailable/NaN: interpolate or use entry price
                    if current_price is None or pd.isna(current_price):
                        # Interpolate price based on elapsed time if we have exit price
                        if not pd.isna(exit_date) and not pd.isna(trade['Exit Price']):
                            total_days = (exit_date - entry_date).days
                            elapsed_days = (date - entry_date).days
                            if total_days > 0:
                                current_price = entry_price + (trade['Exit Price'] - entry_price) * (elapsed_days / total_days)
                            else:
                                current_price = entry_price
                        else:
                            current_price = entry_price
                            
                    total_pnl_sum += qty * (current_price - entry_price)
                    
            total_pnl_df.loc[date, stock] = total_pnl_sum
            
    # Save transposed total P&L race CSV
    total_pnl_df_transposed = total_pnl_df.T
    total_pnl_df_transposed.index.name = "ETF"
    total_pnl_df_transposed.columns = pd.to_datetime(total_pnl_df_transposed.columns).strftime('%d-%b-%Y')
    
    total_pnl_path = "total_pnl_race.csv"
    total_pnl_df_transposed.round(2).to_csv(total_pnl_path)
    print(f"[INFO] Calculated P&L trajectory saved to {total_pnl_path}")
    
    # 7. Prepare D3.js structures for HTML generation
    dates_list = list(total_pnl_df_transposed.columns)
    
    # Map dates to active stock list
    active_trades_dict = {}
    for d_str in dates_list:
        d = pd.to_datetime(d_str, format="%d-%b-%Y")
        active_stocks = []
        for _, trade in df_trades.iterrows():
            if trade['Entry Date'] <= d < trade['Exit Date']:
                active_stocks.append(trade['Stock Name'])
        active_trades_dict[d_str] = active_stocks
        
    # Serialize trades details for JS tables
    trades_list = []
    for _, row in df_trades.iterrows():
        trades_list.append({
            "stock": row["Stock Name"],
            "entryPrice": float(row["Entry Price"]),
            "exitPrice": float(row["Exit Price"]),
            "qty": int(row["Qty"]),
            "pnl": float(row["P&L"]),
            "entryDate": row["Entry Date"].strftime("%d-%b-%Y"),
            "exitDate": row["Exit Date"].strftime("%d-%b-%Y"),
            "exitReason": str(row["Exit Reason"])
        })
        
    # NIFTYBEES prices mapping
    nifty_prices_dict = {}
    if benchmark_ticker in close_df.columns:
        nifty_series = close_df[benchmark_ticker]
        for d in close_df.index:
            n_str = d.strftime("%d-%b-%Y")
            nifty_prices_dict[n_str] = float(nifty_series.loc[d])
    else:
        print("[WARNING] NIFTYBEES.NS prices missing. Using placeholder 200.0.")
        nifty_prices_dict = {d: 200.0 for d in dates_list}
        
    # closing prices dictionary
    prices_dict = {}
    for ticker in close_df.columns:
        clean_key = ticker.replace(".NS", "")
        prices_dict[clean_key] = close_df[ticker].to_dict()
        # Convert keys in price dictionary to strftime
        prices_dict[clean_key] = {
            pd.to_datetime(k).strftime("%d-%b-%Y"): float(v) for k, v in prices_dict[clean_key].items()
        }
        
    # P&L trajectory data mapping
    data_dict = {}
    for d in dates_list:
        data_dict[d] = total_pnl_df_transposed[d].to_dict()
        
    global_max = float(total_pnl_df.max().max())
    global_min = float(total_pnl_df.min().min())
    abs_max_scale = max(abs(global_max), abs(global_min))
    
    # 8. HTML Template Injection
    print("[INFO] Constructing interactive HTML dashboard...")
    
    # Prepare D3.js JSON serializations
    dates_json = json.dumps(dates_list)
    etfs_json = json.dumps(unique_etfs)
    data_json = json.dumps(data_dict)
    active_trades_json = json.dumps(active_trades_dict)
    nifty_prices_json = json.dumps(nifty_prices_dict)
    prices_json = json.dumps(prices_dict)
    trades_json = json.dumps(trades_list)
    
    # Build limit selector options based on ETF count
    limit_options_html = f"""
                    <option value="10">Top 10 ETFs</option>
                    <option value="15">Top 15 ETFs</option>
                    <option value="20">Top 20 ETFs</option>
                    <option value="{len(unique_etfs)}" selected>Show All ({len(unique_etfs)})</option>
    """
    
    # HTML contents from the template
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ETF Strategy P&L Race</title>
    
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    
    <style>
        :root {{
            --bg-color: #0b0f19;
            --panel-bg: #151c2c;
            --text-main: #f3f4f6;
            --text-muted: #9ca3af;
            --accent: #3b82f6;
            --accent-glow: rgba(59, 130, 246, 0.35);
            --green: #10b981;
            --red: #ef4444;
            --grid-line: rgba(255, 255, 255, 0.05);
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            background-color: var(--bg-color);
            color: var(--text-main);
            font-family: 'Outfit', sans-serif;
            padding: 1.5rem;
            height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            overflow: hidden;
        }}

        .app-layout {{
            width: 100%;
            max-width: 1600px;
            height: calc(100vh - 3rem);
            display: flex;
            gap: 1.25rem;
        }}

        .sidebar {{
            width: 260px;
            flex-shrink: 0;
            background-color: var(--panel-bg);
            border-radius: 16px;
            padding: 1.5rem;
            display: flex;
            flex-direction: column;
            gap: 1.25rem;
            border: 1px solid rgba(255, 255, 255, 0.05);
            box-shadow: 0 20px 40px -15px rgba(0, 0, 0, 0.5);
            overflow-y: auto;
        }}

        .title-area h1 {{
            font-size: 1.5rem;
            font-weight: 800;
            line-height: 1.15;
            background: linear-gradient(135deg, #60a5fa, #3b82f6, #1d4ed8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.03em;
        }}

        .title-area p {{
            color: var(--text-muted);
            font-size: 0.78rem;
            margin-top: 0.3rem;
            line-height: 1.4;
        }}

        .stats-panel {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
            background-color: rgba(0, 0, 0, 0.15);
            padding: 1rem;
            border-radius: 10px;
            border: 1px solid rgba(255, 255, 255, 0.03);
        }}

        .stat-box {{
            display: flex;
            flex-direction: column;
            gap: 0.2rem;
        }}

        .stat-label {{
            font-size: 0.7rem;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: var(--text-muted);
            font-weight: 500;
        }}

        .stat-value {{
            font-size: 1.3rem;
            font-weight: 700;
            font-variant-numeric: tabular-nums;
        }}

        #date-display {{
            color: #60a5fa;
            text-shadow: 0 0 10px rgba(96, 165, 250, 0.3);
        }}

        .controls-block {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
        }}

        .button-group {{
            display: flex;
            gap: 0.5rem;
        }}

        button {{
            flex: 1;
            background-color: rgba(255, 255, 255, 0.06);
            border: 1px solid rgba(255, 255, 255, 0.08);
            color: var(--text-main);
            padding: 0.5rem;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-family: inherit;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.3rem;
            transition: all 0.2s ease;
            font-size: 0.8rem;
        }}

        button:hover {{
            background-color: var(--accent);
            border-color: var(--accent);
            box-shadow: 0 0 12px var(--accent-glow);
        }}

        button:active {{
            transform: scale(0.97);
        }}

        .control-item {{
            display: flex;
            flex-direction: column;
            gap: 0.3rem;
        }}

        .control-item label {{
            font-size: 0.65rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            font-weight: 600;
        }}

        select, input[type="range"] {{
            background-color: rgba(0, 0, 0, 0.2);
            border: 1px solid rgba(255, 255, 255, 0.1);
            color: var(--text-main);
            padding: 0.4rem 0.5rem;
            border-radius: 6px;
            font-family: inherit;
            font-weight: 500;
            outline: none;
            cursor: pointer;
            font-size: 0.8rem;
            width: 100%;
        }}

        select:focus {{
            border-color: var(--accent);
        }}

        input[type="range"] {{
            height: 6px;
            -webkit-appearance: none;
            background: rgba(255, 255, 255, 0.1);
            border: none;
            border-radius: 3px;
        }}

        input[type="range"]::-webkit-slider-thumb {{
            -webkit-appearance: none;
            width: 14px;
            height: 14px;
            border-radius: 50%;
            background: var(--accent);
            cursor: pointer;
            box-shadow: 0 0 5px var(--accent-glow);
        }}

        .middle-column {{
            flex-grow: 1;
            display: flex;
            flex-direction: column;
            gap: 1.25rem;
            height: 100%;
            overflow: hidden;
        }}

        .scrubber-container {{
            display: flex;
            align-items: center;
            gap: 1rem;
            background-color: var(--panel-bg);
            padding: 0.75rem 1.25rem;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.05);
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.2);
            flex-shrink: 0;
        }}

        #timeline-scrubber {{
            flex-grow: 1;
        }}

        .chart-wrapper {{
            background-color: var(--panel-bg);
            border-radius: 16px;
            padding: 1.75rem 2.25rem 2.25rem 2.25rem;
            box-shadow: 0 20px 40px -15px rgba(0, 0, 0, 0.5);
            border: 1px solid rgba(255, 255, 255, 0.05);
            flex-grow: 1; 
            display: flex;
            flex-direction: column;
            overflow: hidden;
            position: relative;
        }}

        .chart-area {{
            position: relative;
            width: 100%;
            height: 100%;
            transition: height 0.4s ease;
        }}

        .grid-lines {{
            position: absolute;
            top: 0;
            bottom: 0;
            left: 180px;
            right: 140px;
            pointer-events: none;
            display: flex;
            justify-content: space-between;
            z-index: 1;
            transition: left 0.4s ease, right 0.4s ease;
        }}

        .grid-line {{
            width: 1px;
            height: 100%;
            background-color: var(--grid-line);
            position: relative;
        }}

        .grid-label {{
            position: absolute;
            bottom: -22px;
            transform: translateX(-50%);
            font-size: 0.7rem;
            color: var(--text-muted);
            font-weight: 500;
            white-space: nowrap;
        }}

        .bar-row {{
            position: absolute;
            left: 0;
            height: 38px;
            width: 100%;
            display: flex;
            align-items: center;
            z-index: 2;
        }}

        .bar-row.active {{
            opacity: 1.0;
        }}

        .bar-row.inactive {{
            opacity: 0.35;
        }}

        .active-indicator {{
            width: 6px;
            height: 6px;
            background-color: var(--green);
            border-radius: 50%;
            display: inline-block;
            box-shadow: 0 0 8px var(--green);
            margin-left: 4px;
            animation: pulse 1.5s infinite;
        }}

        @keyframes pulse {{
            0% {{ transform: scale(0.9); opacity: 0.6; }}
            50% {{ transform: scale(1.25); opacity: 1; }}
            100% {{ transform: scale(0.9); opacity: 0.6; }}
        }}

        .bar-label {{
            width: 160px;
            font-size: 0.85rem;
            font-weight: 700;
            color: var(--text-main);
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            padding-right: 12px;
            text-align: right;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            gap: 0.4rem;
            transition: transform 0.2s ease, color 0.2s ease, width 0.4s ease, font-size 0.4s ease;
        }}

        .bar-track {{
            flex-grow: 1;
            height: 100%;
            display: flex;
            align-items: center;
            position: relative;
        }}

        .bar-fill {{
            height: 20px;
            border-radius: 4px;
            position: absolute;
            top: auto;
            left: 0;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            box-shadow: 0 4px 10px rgba(0, 0, 0, 0.15);
        }}

        .bar-row:hover .bar-fill {{
            filter: brightness(1.2);
            transform: scaleY(1.08);
            cursor: pointer;
        }}
        
        .bar-row:hover .bar-label {{
            transform: translateX(3px);
            color: #ffffff;
        }}

        .bar-value {{
            width: 120px;
            font-size: 0.82rem;
            font-weight: 700;
            font-variant-numeric: tabular-nums;
            padding-left: 12px;
            text-align: left;
            transition: color 0.2s ease, font-size 0.4s ease, width 0.4s ease;
            overflow: hidden;
            text-overflow: clip;
            white-space: nowrap;
        }}

        .bar-value.positive {{
            color: var(--green);
        }}

        .bar-value.negative {{
            color: var(--red);
        }}

        .bottom-stats-panel {{
            background-color: var(--panel-bg);
            border-radius: 12px;
            padding: 0.85rem 1.5rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border: 1px solid rgba(255, 255, 255, 0.05);
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.2);
            gap: 1.25rem;
            flex-shrink: 0;
        }}

        .bottom-stat-box {{
            display: flex;
            flex-direction: column;
            gap: 0.15rem;
            flex: 1;
            align-items: center;
            min-width: 0;
        }}

        .bottom-stat-label {{
            font-size: 0.68rem;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: var(--text-muted);
            font-weight: 600;
            white-space: nowrap;
        }}

        .bottom-stat-value-neutral {{
            font-size: 1.35rem;
            font-weight: 800;
            color: var(--text-main);
            font-variant-numeric: tabular-nums;
            white-space: nowrap;
        }}

        .bottom-stat-value {{
            font-size: 1.35rem;
            font-weight: 800;
            font-variant-numeric: tabular-nums;
            transition: color 0.2s ease, font-size 0.2s ease;
            white-space: nowrap;
        }}

        .stat-divider {{
            width: 1px;
            height: 30px;
            background-color: rgba(255, 255, 255, 0.1);
            flex-shrink: 0;
        }}

        .bottom-stat-value.positive {{
            color: var(--green);
            text-shadow: 0 0 10px rgba(16, 185, 129, 0.25);
        }}

        .bottom-stat-value.negative {{
            color: var(--red);
            text-shadow: 0 0 10px rgba(239, 68, 68, 0.25);
        }}

        .right-column {{
            width: 440px;
            flex-shrink: 0;
            display: flex;
            flex-direction: column;
            gap: 0.85rem;
            height: 100%;
            overflow: hidden;
        }}

        .benchmark-panel {{
            background-color: rgba(0, 0, 0, 0.2);
            border: 1px dashed rgba(255, 255, 255, 0.12);
            border-radius: 12px;
            padding: 0.85rem;
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
            flex-shrink: 0;
        }}

        .benchmark-header {{
            display: flex;
            flex-direction: column;
            border-bottom: 1px solid rgba(255, 255, 255, 0.08);
            padding-bottom: 0.4rem;
        }}

        .benchmark-title {{
            font-size: 0.78rem;
            font-weight: 800;
            color: #60a5fa;
            text-transform: uppercase;
            letter-spacing: 0.06em;
        }}

        .benchmark-subtitle {{
            font-size: 0.65rem;
            color: var(--text-muted);
            font-weight: 500;
        }}

        .benchmark-content {{
            display: flex;
            flex-direction: column;
            gap: 0.3rem;
        }}

        .benchmark-row {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-size: 0.75rem;
        }}

        .benchmark-label {{
            color: var(--text-muted);
            font-weight: 500;
        }}

        .benchmark-val {{
            font-weight: 700;
            font-variant-numeric: tabular-nums;
            color: var(--text-main);
            transition: font-size 0.2s ease;
        }}

        .benchmark-val.positive {{
            color: var(--green);
        }}

        .benchmark-val.negative {{
            color: var(--red);
        }}

        .sim-panel {{
            background-color: var(--panel-bg);
            border-radius: 12px;
            padding: 0.85rem;
            border: 1px solid rgba(255, 255, 255, 0.05);
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.35);
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
            min-height: 0;
        }}

        .sim-panel-title {{
            font-size: 0.78rem;
            font-weight: 700;
            color: #60a5fa;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            border-bottom: 1px solid rgba(255, 255, 255, 0.08);
            padding-bottom: 0.35rem;
            flex-shrink: 0;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .table-container {{
            flex-grow: 1;
            overflow-y: auto;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.72rem;
            text-align: left;
        }}

        th {{
            position: sticky;
            top: 0;
            background-color: var(--panel-bg);
            color: var(--text-muted);
            font-weight: 600;
            padding: 0.35rem;
            font-size: 0.65rem;
            text-transform: uppercase;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
            z-index: 10;
        }}

        td {{
            padding: 0.3rem 0.35rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.03);
            font-variant-numeric: tabular-nums;
        }}

        tr:hover td {{
            background-color: rgba(255, 255, 255, 0.02);
        }}

        .text-green {{
            color: var(--green) !important;
            font-weight: 600;
        }}

        .text-red {{
            color: var(--red) !important;
            font-weight: 600;
        }}

        .empty-row-msg {{
            color: var(--text-muted);
            text-align: center;
            padding: 1rem;
            font-style: italic;
            font-size: 0.72rem;
        }}

        @media (max-width: 1100px) {{
            body {{
                overflow-y: auto;
                height: auto;
                padding: 1rem;
            }}
            .app-layout {{
                flex-direction: column;
                height: auto;
                gap: 1.25rem;
            }}
            .sidebar {{
                width: 100%;
                height: auto;
                overflow-y: visible;
            }}
            .middle-column {{
                height: 500px;
                flex-grow: 0;
            }}
            .right-column {{
                width: 100%;
                height: auto;
                overflow-y: visible;
                gap: 1.25rem;
            }}
            .sim-panel {{
                height: 250px;
            }}
        }}

        .hidden {{
            display: none !important;
        }}

        svg {{
            width: 16px;
            height: 16px;
            fill: currentColor;
        }}
    </style>
</head>
<body>

<div class="app-layout">
    <aside class="sidebar">
        <div class="title-area">
            <h1>Momentify<br>ETF Strategy</h1>
            <p>Interactive timeline tracking cumulative realized & unrealized performance</p>
        </div>
        
        <div class="stats-panel">
            <div class="stat-box">
                <div class="stat-label">Start Date</div>
                <div class="stat-value" id="start-date-display" style="color: var(--text-main); font-size: 1.3rem;">--/--/----</div>
            </div>
            <div class="stat-box">
                <div class="stat-label">Current Date</div>
                <div class="stat-value" id="date-display" style="font-size: 1.3rem;">--/--/----</div>
            </div>
        </div>

        <div class="controls-block">
            <div class="button-group">
                <button id="play-btn">
                    <svg id="play-icon" viewBox="0 0 24 24"><path d="M8 5v14l11-7z"/></svg>
                    <svg id="pause-icon" class="hidden" viewBox="0 0 24 24"><path d="M6 19h4V5H6v14zm8-14v14h4V5h-4z"/></svg>
                    <span id="play-text">Play</span>
                </button>
                <button id="reset-btn">
                    <svg viewBox="0 0 24 24"><path d="M12 5V1L7 6l5 5V7c3.31 0 6 2.69 6 6s-2.69 6-6 6-6-2.69-6-6H4c0 4.42 3.58 8 8 8s8-3.58 8-8-3.58-8-8-8z"/></svg>
                    <span>Reset</span>
                </button>
            </div>

            <div class="control-item">
                <label for="speed-range">Animation Speed</label>
                <input type="range" id="speed-range" min="100" max="1000" step="50" value="700">
            </div>

            <div class="control-item">
                <label for="limit-select">Show Top</label>
                <select id="limit-select">
                    {limit_options_html}
                </select>
            </div>
        </div>
    </aside>

    <main class="middle-column">
        <div class="bottom-stats-panel">
            <div class="bottom-stat-box">
                <span class="bottom-stat-label">Invested Capital</span>
                <span class="bottom-stat-value-neutral">₹5,00,000.00</span>
            </div>
            <div class="stat-divider"></div>
            <div class="bottom-stat-box">
                <span class="bottom-stat-label">Total Portfolio P&L</span>
                <span class="bottom-stat-value" id="total-pnl-display">₹0.00</span>
            </div>
            <div class="stat-divider"></div>
            <div class="bottom-stat-box">
                <span class="bottom-stat-label">Absolute Return</span>
                <span class="bottom-stat-value" id="portfolio-pct-display">0.00%</span>
            </div>
            <div class="stat-divider"></div>
            <div class="bottom-stat-box">
                <span class="bottom-stat-label">System CAGR</span>
                <span class="bottom-stat-value" id="portfolio-cagr-display">0.00% p.a.</span>
            </div>
            <div class="stat-divider"></div>
            <div class="bottom-stat-box">
                <span class="bottom-stat-label">Max Drawdown</span>
                <span class="bottom-stat-value" id="portfolio-maxdd-display" style="color: var(--red);">0.00%</span>
            </div>
        </div>

        <div class="chart-wrapper">
            <div class="chart-area" id="chart-container">
                <div class="grid-lines" id="grid-lines-container"></div>
            </div>
        </div>

        <div class="scrubber-container">
            <span id="start-date-lbl" style="font-size: 0.75rem; color: var(--text-muted); font-weight:600;">Start</span>
            <input type="range" id="timeline-scrubber" min="0" max="0" value="0">
            <span id="end-date-lbl" style="font-size: 0.75rem; color: var(--text-muted); font-weight:600;">End</span>
        </div>
    </main>

    <aside class="right-column">
        <div class="benchmark-panel">
            <div class="benchmark-header">
                <span class="benchmark-title">NIFTYBEES Buy & Hold</span>
                <span class="benchmark-subtitle">Alternative 5L Investment</span>
            </div>
            <div class="benchmark-content">
                <div class="benchmark-row">
                    <span class="benchmark-label">Value:</span>
                    <span class="benchmark-val" id="nifty-val-display">₹5,00,000.00</span>
                </div>
                <div class="benchmark-row">
                    <span class="benchmark-label">P&L:</span>
                    <span class="benchmark-val" id="nifty-pnl-display">₹0.00</span>
                </div>
                <div class="benchmark-row">
                    <span class="benchmark-label">Return:</span>
                    <span class="benchmark-val" id="nifty-pct-display">0.00%</span>
                </div>
                <div class="benchmark-row">
                    <span class="benchmark-label">CAGR:</span>
                    <span class="benchmark-val" id="nifty-cagr-display">0.00% p.a.</span>
                </div>
                <div class="benchmark-row">
                    <span class="benchmark-label">Max Drawdown:</span>
                    <span class="benchmark-val" id="nifty-maxdd-display" style="color: var(--red);">0.00%</span>
                </div>
            </div>
        </div>

        <div class="sim-panel" style="flex: 2;">
            <div class="sim-panel-title">Running Stocks (Open Positions)</div>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Stock</th>
                            <th>Entry Price</th>
                            <th>Qty</th>
                            <th>Current Price</th>
                            <th>P&L</th>
                            <th>Return</th>
                        </tr>
                    </thead>
                    <tbody id="running-body"></tbody>
                </table>
            </div>
        </div>

        <div class="sim-panel" style="flex: 1.2;">
            <div class="sim-panel-title" id="exited-title">Exited Stocks (as of --/--/----)</div>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Stock</th>
                            <th>Entry</th>
                            <th>Exit</th>
                            <th>Qty</th>
                            <th>P&L</th>
                            <th>Return</th>
                        </tr>
                    </thead>
                    <tbody id="exited-body"></tbody>
                </table>
            </div>
        </div>

        <div class="sim-panel" style="flex: 1.2;">
            <div class="sim-panel-title" id="added-title">Added Stocks (as of --/--/----)</div>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Stock</th>
                            <th>Entry Price</th>
                            <th>Qty</th>
                            <th>Amount</th>
                        </tr>
                    </thead>
                    <tbody id="added-body"></tbody>
                </table>
            </div>
        </div>
    </aside>
</div>

<script>
    // Embedded Data
    const dates = {dates_json};
    const etfs = {etfs_json};
    const dataset = {data_json};
    const activeTrades = {active_trades_json};
    const niftyPrices = {nifty_prices_json};
    const prices = {prices_json};
    const tradesList = {trades_json};
    const globalMaxAbs = {abs_max_scale};

    // App State
    let currentIndex = 0;
    let isPlaying = false;
    let timer = null;
    let speed = 400; 
    let topN = {len(unique_etfs)}; // show all
    const scaleMode = 'absolute'; 

    // Colors mapping (strictly Green-based for positive, Red-based for negative)
    const colorMap = {{}};
    etfs.forEach(etf => {{
        let hash = 0;
        for (let i = 0; i < etf.length; i++) {{
            hash = etf.charCodeAt(i) + ((hash << 5) - hash);
        }}
        const greenHue = 120 + Math.abs(hash % 40);
        const redHue = (345 + Math.abs(hash % 20)) % 360;

        colorMap[etf] = {{
            greenBase: `hsl(${{greenHue}}, 70%, 42%)`,
            greenLight: `hsl(${{greenHue}}, 80%, 56%)`,
            greenGlow: `rgba(16, 185, 129, 0.35)`,
            redBase: `hsl(${{redHue}}, 70%, 42%)`,
            redLight: `hsl(${{redHue}}, 80%, 56%)`,
            redGlow: `rgba(239, 68, 68, 0.35)`
        }};
    }});

    // DOM Elements
    const playBtn = document.getElementById('play-btn');
    const playText = document.getElementById('play-text');
    const playIcon = document.getElementById('play-icon');
    const pauseIcon = document.getElementById('pause-icon');
    const resetBtn = document.getElementById('reset-btn');
    const speedRange = document.getElementById('speed-range');
    const limitSelect = document.getElementById('limit-select');
    const timelineScrubber = document.getElementById('timeline-scrubber');
    const dateDisplay = document.getElementById('date-display');
    const startDisplay = document.getElementById('start-date-display');
    const totalPnlDisplay = document.getElementById('total-pnl-display');
    const portfolioPctDisplay = document.getElementById('portfolio-pct-display');
    const portfolioCagrDisplay = document.getElementById('portfolio-cagr-display');
    const portfolioMaxDdDisplay = document.getElementById('portfolio-maxdd-display');
    const chartContainer = document.getElementById('chart-container');
    const gridContainer = document.getElementById('grid-lines-container');
    const startDateLbl = document.getElementById('start-date-lbl');
    const endDateLbl = document.getElementById('end-date-lbl');

    // Nifty Benchmark DOM
    const niftyValDisplay = document.getElementById('nifty-val-display');
    const niftyPnlDisplay = document.getElementById('nifty-pnl-display');
    const niftyPctDisplay = document.getElementById('nifty-pct-display');
    const niftyCagrDisplay = document.getElementById('nifty-cagr-display');
    const niftyMaxDdDisplay = document.getElementById('nifty-maxdd-display');

    // Simulator Tables DOM
    const runningBody = document.getElementById('running-body');
    const exitedBody = document.getElementById('exited-body');
    const addedBody = document.getElementById('added-body');
    const exitedTitle = document.getElementById('exited-title');
    const addedTitle = document.getElementById('added-title');

    // Initialize App
    function init() {{
        timelineScrubber.max = dates.length - 1;
        startDisplay.textContent = dates[0];
        startDateLbl.textContent = dates[0];
        endDateLbl.textContent = dates[dates.length - 1];

        // Parse speed from slider
        speed = 1100 - parseInt(speedRange.value);

        createBarElements();
        updateFrame();

        playBtn.addEventListener('click', togglePlay);
        resetBtn.addEventListener('click', resetChart);
        speedRange.addEventListener('input', updateSpeed);
        limitSelect.addEventListener('change', updateLimit);
        timelineScrubber.addEventListener('input', scrubTimeline);

        window.addEventListener('resize', () => {{
            updateFrame();
        }});
    }}

    function createBarElements() {{
        const gridHtml = gridContainer.outerHTML;
        chartContainer.innerHTML = gridHtml;
        gridContainer_ref = document.getElementById('grid-lines-container');

        etfs.forEach(etf => {{
            const row = document.createElement('div');
            row.className = 'bar-row';
            row.id = `row-${{etf}}`;
            row.style.top = '1200px';

            const label = document.createElement('div');
            label.className = 'bar-label';
            label.id = `lbl-${{etf}}`;
            label.innerHTML = `<span style="font-weight:600;">${{etf}}</span>`;

            const track = document.createElement('div');
            track.className = 'bar-track';

            const fill = document.createElement('div');
            fill.className = 'bar-fill positive';
            fill.id = `fill-${{etf}}`;
            fill.style.width = '0%';
            fill.style.backgroundColor = colorMap[etf].greenBase;
            fill.style.boxShadow = `0 0 10px ${{colorMap[etf].greenGlow}}`;

            track.appendChild(fill);

            const value = document.createElement('div');
            value.className = 'bar-value';
            value.id = `val-${{etf}}`;
            value.textContent = '₹0.00';

            row.appendChild(label);
            row.appendChild(track);
            row.appendChild(value);
            chartContainer.appendChild(row);
        }});
    }}

    function setAdaptiveFontRem(element, text, baseRem) {{
        element.textContent = text;
        if (text.length > 15) {{
            element.style.fontSize = (baseRem * 0.72) + 'rem';
        }} else if (text.length > 12) {{
            element.style.fontSize = (baseRem * 0.84) + 'rem';
        }} else {{
            element.style.fontSize = baseRem + 'rem';
        }}
    }}

    function updateFrame() {{
        const currentDate = dates[currentIndex];
        dateDisplay.textContent = currentDate;
        timelineScrubber.value = currentIndex;

        exitedTitle.textContent = `Exited Stocks (as of ${{currentDate}})`;
        addedTitle.textContent = `Added Stocks (as of ${{currentDate}})`;

        const dateData = dataset[currentDate];
        const activeList = activeTrades[currentDate] || [];

        const containerWidth = chartContainer.clientWidth;
        const containerHeight = chartContainer.clientHeight || 500; 

        let labelWidth = 140;
        let valueWidth = 115;
        if (containerWidth < 700) {{
            labelWidth = 90;
            valueWidth = 90;
        }} else if (containerWidth > 1100) {{
            labelWidth = 160;
            valueWidth = 125;
        }}

        const rowOffset = containerHeight / topN;
        const rowHeight = Math.max(8, rowOffset - 3);
        const barHeight = Math.min(28, Math.max(5, rowHeight * 0.65));
        const fontSize = Math.min(16, Math.max(9, rowHeight * 0.55)) + 'px';

        const gridContainer_ref = document.getElementById('grid-lines-container');
        if (gridContainer_ref) {{
            gridContainer_ref.style.left = `${{labelWidth}}px`;
            gridContainer_ref.style.right = `${{valueWidth}}px`;
        }}

        const startDate = new Date(dates[0].replace(/-/g, ' '));
        const currentDateObj = new Date(currentDate.replace(/-/g, ' '));
        const diffTime = Math.abs(currentDateObj - startDate);
        const diffDays = Math.ceil(diffTime / (1000 * 60 * 60 * 24));
        const diffYears = diffDays / 365.25;

        let totalPnl = 0.0;
        etfs.forEach(etf => {{
            totalPnl += dateData[etf] || 0.0;
        }});

        const formattedTotal = formatCurrency(totalPnl);
        setAdaptiveFontRem(totalPnlDisplay, formattedTotal, 1.35);
        if (totalPnl >= 0) {{
            totalPnlDisplay.className = 'bottom-stat-value positive';
        }} else {{
            totalPnlDisplay.className = 'bottom-stat-value negative';
        }}

        const portfolioPct = (totalPnl / 500000.0) * 100;
        const pctText = (portfolioPct >= 0 ? '+' : '') + portfolioPct.toFixed(2) + '%';
        setAdaptiveFontRem(portfolioPctDisplay, pctText, 1.35);
        portfolioPctDisplay.className = 'bottom-stat-value ' + (portfolioPct >= 0 ? 'positive' : 'negative');

        let portfolioCagr = 'N/A';
        if (diffYears >= 0.1) {{
            const cagrVal = (Math.pow(1 + (totalPnl / 500000.0), 1 / diffYears) - 1) * 100;
            portfolioCagr = cagrVal.toFixed(2) + '% p.a.';
            setAdaptiveFontRem(portfolioCagrDisplay, portfolioCagr, 1.35);
            portfolioCagrDisplay.className = 'bottom-stat-value ' + (portfolioPct >= 0 ? 'positive' : 'negative');
        }} else {{
            portfolioCagrDisplay.textContent = 'N/A';
            portfolioCagrDisplay.style.fontSize = '1.35rem';
            portfolioCagrDisplay.className = 'bottom-stat-value';
        }}

        let peakValue = 500000.0;
        let maxDrawdown = 0.0;
        for (let i = 0; i <= currentIndex; i++) {{
            const date = dates[i];
            let stepPnl = 0.0;
            etfs.forEach(etf => {{
                stepPnl += dataset[date][etf] || 0.0;
            }});
            const portfolioValue = 500000.0 + stepPnl;
            if (portfolioValue > peakValue) {{
                peakValue = portfolioValue;
            }}
            const dd = ((peakValue - portfolioValue) / peakValue) * 100.0;
            if (dd > maxDrawdown) {{
                maxDrawdown = dd;
            }}
        }}
        const maxDdText = '-' + maxDrawdown.toFixed(2) + '%';
        setAdaptiveFontRem(portfolioMaxDdDisplay, maxDdText, 1.35);

        const niftyStartPrice = niftyPrices[dates[0]];
        const niftyCurrentPrice = niftyPrices[currentDate];
        let niftyVal = 500000.0;
        let niftyPnl = 0.0;
        let niftyPct = 0.0;
        let niftyCagr = 'N/A';

        if (niftyStartPrice && niftyCurrentPrice) {{
            const priceRatio = niftyCurrentPrice / niftyStartPrice;
            niftyVal = 500000.0 * priceRatio;
            niftyPnl = niftyVal - 500000.0;
            niftyPct = (priceRatio - 1.0) * 100;

            if (diffYears >= 0.1) {{
                const niftyCagrVal = (Math.pow(priceRatio, 1 / diffYears) - 1) * 100;
                niftyCagr = niftyCagrVal.toFixed(2) + '% p.a.';
            }}
        }}

        let niftyPeak = 500000.0;
        let niftyMaxDD = 0.0;
        for (let i = 0; i <= currentIndex; i++) {{
            const date = dates[i];
            const price = niftyPrices[date];
            if (price && niftyStartPrice) {{
                const val = 500000.0 * (price / niftyStartPrice);
                if (val > niftyPeak) {{
                    niftyPeak = val;
                }}
                const dd = ((niftyPeak - val) / niftyPeak) * 100.0;
                if (dd > niftyMaxDD) {{
                    niftyMaxDD = dd;
                }}
            }}
        }}

        niftyValDisplay.textContent = '₹' + niftyVal.toLocaleString('en-IN', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
        niftyPnlDisplay.textContent = (niftyPnl >= 0 ? '+' : '') + '₹' + Math.abs(niftyPnl).toLocaleString('en-IN', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
        niftyPnlDisplay.className = 'benchmark-val ' + (niftyPnl >= 0 ? 'positive' : 'negative');
        
        niftyPctDisplay.textContent = (niftyPct >= 0 ? '+' : '') + niftyPct.toFixed(2) + '%';
        niftyPctDisplay.className = 'benchmark-val ' + (niftyPct >= 0 ? 'positive' : 'negative');

        niftyCagrDisplay.textContent = niftyCagr === 'N/A' ? 'N/A' : niftyCagr;
        niftyCagrDisplay.className = 'benchmark-val ' + (niftyPct >= 0 ? 'positive' : 'negative');

        niftyMaxDdDisplay.textContent = '-' + niftyMaxDD.toFixed(2) + '%';

        populateActivityTables(currentDate);

        const items = etfs.map(etf => ({{
            etf: etf,
            val: dateData[etf] || 0.0
        }}));

        items.sort((a, b) => b.val - a.val);
        const maxVal = globalMaxAbs;

        drawGrid(maxVal, labelWidth, valueWidth);

        const duration = isPlaying ? speed : 200;

        items.forEach((item, index) => {{
            const row = document.getElementById(`row-${{item.etf}}`);
            const fill = document.getElementById(`fill-${{item.etf}}`);
            const valDiv = document.getElementById(`val-${{item.etf}}`);
            const labelDiv = document.getElementById(`lbl-${{item.etf}}`);

            row.style.transition = `top ${{duration}}ms cubic-bezier(0.25, 0.1, 0.25, 1), opacity ${{duration}}ms ease`;
            fill.style.transition = `width ${{duration}}ms cubic-bezier(0.25, 0.1, 0.25, 1)`;

            if (index < topN) {{
                row.style.top = `${{index * rowOffset}}px`;
                row.style.height = `${{rowHeight}}px`;
                fill.style.height = `${{barHeight}}px`;
                labelDiv.style.width = labelWidth + 'px';
                labelDiv.style.fontSize = fontSize;
                valDiv.style.width = valueWidth + 'px';
                
                row.style.pointerEvents = 'auto';

                const isActive = activeList.includes(item.etf);
                if (isActive) {{
                    row.style.opacity = '1.0';
                    if (!labelDiv.querySelector('.active-indicator')) {{
                        labelDiv.innerHTML = `<span style="font-weight:600;">${{item.etf}}</span><span class="active-indicator"></span>`;
                    }}
                }} else {{
                    row.style.opacity = '0.35';
                    labelDiv.innerHTML = `<span style="font-weight:600;">${{item.etf}}</span>`;
                }}

                let pct = (Math.abs(item.val) / maxVal) * 100;
                if (pct < 1 && Math.abs(item.val) > 0) pct = 1;

                fill.style.width = `${{pct}}%`;

                if (item.val >= 0) {{
                    fill.style.background = `linear-gradient(90deg, ${{colorMap[item.etf].greenBase}}, ${{colorMap[item.etf].greenLight}})`;
                    fill.style.boxShadow = isActive ? `0 0 10px ${{colorMap[item.etf].greenGlow}}` : 'none';
                    valDiv.className = 'bar-value positive';
                }} else {{
                    fill.style.background = `linear-gradient(90deg, ${{colorMap[item.etf].redBase}}, ${{colorMap[item.etf].redLight}})`;
                    fill.style.boxShadow = isActive ? `0 0 10px ${{colorMap[item.etf].redGlow}}` : 'none';
                    valDiv.className = 'bar-value negative';
                }}

                const valText = formatCurrency(item.val);
                valDiv.textContent = valText;

                const basePx = parseFloat(fontSize);
                let adjustedFontSize = fontSize;
                if (valText.length > 12) {{
                    adjustedFontSize = (basePx * 0.74) + 'px';
                }} else if (valText.length > 9) {{
                    adjustedFontSize = (basePx * 0.86) + 'px';
                }}
                valDiv.style.fontSize = adjustedFontSize;

            }} else {{
                row.style.top = `${{topN * rowOffset}}px`;
                row.style.opacity = '0';
                row.style.pointerEvents = 'none';
            }}
        }});
    }}

    function populateActivityTables(currentDate) {{
        runningBody.innerHTML = '';
        exitedBody.innerHTML = '';
        addedBody.innerHTML = '';

        const currentIdx = dates.indexOf(currentDate);

        const runningTrades = [];
        const exitedTrades = [];
        const addedTrades = [];

        tradesList.forEach(trade => {{
            const entryIdx = dates.indexOf(trade.entryDate);
            const exitIdx = dates.indexOf(trade.exitDate);

            if (entryIdx !== -1 && exitIdx !== -1 && entryIdx <= currentIdx && exitIdx > currentIdx) {{
                const currentPrice = prices[trade.stock] ? prices[trade.stock][currentDate] : null;
                const finalPrice = currentPrice !== null ? currentPrice : trade.entryPrice;
                const pnl = (finalPrice - trade.entryPrice) * trade.qty;
                const pct = ((finalPrice / trade.entryPrice) - 1.0) * 100.0;

                runningTrades.push({{
                    stock: trade.stock,
                    entryPrice: trade.entryPrice,
                    qty: trade.qty,
                    currentPrice: finalPrice,
                    pnl: pnl,
                    pct: pct
                }});
            }}

            if (trade.exitDate === currentDate) {{
                const pct = ((trade.exitPrice / trade.entryPrice) - 1.0) * 100.0;
                exitedTrades.push({{
                    stock: trade.stock,
                    entryPrice: trade.entryPrice,
                    exitPrice: trade.exitPrice,
                    qty: trade.qty,
                    pnl: trade.pnl,
                    pct: pct
                }});
            }}

            if (trade.entryDate === currentDate) {{
                addedTrades.push({{
                    stock: trade.stock,
                    entryPrice: trade.entryPrice,
                    qty: trade.qty,
                    amount: trade.entryPrice * trade.qty
                }});
            }}
        }});

        runningTrades.sort((a, b) => b.pnl - a.pnl);

        if (runningTrades.length === 0) {{
            runningBody.innerHTML = `<tr><td colspan="6" class="empty-row-msg">No active open positions on this date</td></tr>`;
        }} else {{
            runningTrades.forEach((t, i) => {{
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td style="font-weight:700; color: #60a5fa;">${{t.stock}}</td>
                    <td>₹${{t.entryPrice.toLocaleString('en-IN', {{minimumFractionDigits: 1}})}}</td>
                    <td>${{t.qty}}</td>
                    <td>₹${{t.currentPrice.toLocaleString('en-IN', {{minimumFractionDigits: 1}})}}</td>
                    <td class="${{t.pnl >= 0 ? 'text-green' : 'text-red'}}">${{t.pnl >= 0 ? '+' : ''}}₹${{Math.round(t.pnl).toLocaleString('en-IN')}}</td>
                    <td class="${{t.pct >= 0 ? 'text-green' : 'text-red'}}">${{t.pct >= 0 ? '+' : ''}}${{t.pct.toFixed(2)}}%</td>
                `;
                runningBody.appendChild(row);
            }});
        }}

        if (exitedTrades.length === 0) {{
            exitedBody.innerHTML = `<tr><td colspan="6" class="empty-row-msg">No closed positions on this date</td></tr>`;
        }} else {{
            exitedTrades.forEach((t, i) => {{
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td style="font-weight:700; color: #60a5fa;">${{t.stock}}</td>
                    <td>₹${{t.entryPrice.toLocaleString('en-IN', {{minimumFractionDigits: 1}})}}</td>
                    <td>₹${{t.exitPrice.toLocaleString('en-IN', {{minimumFractionDigits: 1}})}}</td>
                    <td>${{t.qty}}</td>
                    <td class="${{t.pnl >= 0 ? 'text-green' : 'text-red'}}">${{t.pnl >= 0 ? '+' : ''}}₹${{Math.round(t.pnl).toLocaleString('en-IN')}}</td>
                    <td class="${{t.pct >= 0 ? 'text-green' : 'text-red'}}">${{t.pct >= 0 ? '+' : ''}}${{t.pct.toFixed(2)}}%</td>
                `;
                exitedBody.appendChild(row);
            }});
        }}

        if (addedTrades.length === 0) {{
            addedBody.innerHTML = `<tr><td colspan="4" class="empty-row-msg">No new positions opened on this date</td></tr>`;
        }} else {{
            addedTrades.forEach((t, i) => {{
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td style="font-weight:700; color: #60a5fa;">${{t.stock}}</td>
                    <td>₹${{t.entryPrice.toLocaleString('en-IN', {{minimumFractionDigits: 1}})}}</td>
                    <td>${{t.qty}}</td>
                    <td>₹${{Math.round(t.amount).toLocaleString('en-IN')}}</td>
                `;
                addedBody.appendChild(row);
            }});
        }}
    }}

    function drawGrid(maxVal, labelWidth, valueWidth) {{
        const gridContainer_ref = document.getElementById('grid-lines-container');
        if (!gridContainer_ref) return;
        
        gridContainer_ref.innerHTML = '';
        const steps = 4;
        
        const duration = isPlaying ? speed : 200;
        gridContainer_ref.style.transition = `left ${{duration}}ms ease, right ${{duration}}ms ease`;
        gridContainer_ref.style.left = `${{labelWidth}}px`;
        gridContainer_ref.style.right = `${{valueWidth}}px`;

        for (let i = 0; i <= steps; i++) {{
            const pct = (i / steps) * 100;
            const val = (i / steps) * maxVal;

            const gridLine = document.createElement('div');
            gridLine.className = 'grid-line';
            gridLine.style.left = `${{pct}}%`;

            const label = document.createElement('div');
            label.className = 'grid-label';
            label.textContent = formatCurrencyShort(val);

            gridLine.appendChild(label);
            gridContainer_ref.appendChild(gridLine);
        }}
    }}

    function formatCurrency(val) {{
        const sign = val >= 0 ? '+' : '-';
        return `${{sign}}₹${{Math.abs(val).toLocaleString('en-IN', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}}`;
    }}

    function formatCurrencyShort(val) {{
        const absVal = Math.abs(val);
        if (absVal >= 100000) {{
            return `₹${{(absVal / 100000).toFixed(1)}}L`;
        }} else if (absVal >= 1000) {{
            return `₹${{(absVal / 1000).toFixed(0)}}K`;
        }}
        return `₹${{absVal.toFixed(0)}}`;
    }}

    function togglePlay() {{
        if (isPlaying) {{
            pause();
        }} else {{
            play();
        }}
    }}

    function play() {{
        isPlaying = true;
        playText.textContent = 'Pause';
        playIcon.classList.add('hidden');
        pauseIcon.classList.remove('hidden');

        timer = setInterval(() => {{
            if (currentIndex < dates.length - 1) {{
                currentIndex++;
                updateFrame();
            }} else {{
                pause();
            }}
        }}, speed);
    }}

    function pause() {{
        isPlaying = false;
        playText.textContent = 'Play';
        playIcon.classList.remove('hidden');
        pauseIcon.classList.add('hidden');
        clearInterval(timer);
    }}

    function resetChart() {{
        pause();
        currentIndex = 0;
        updateFrame();
    }}

    function updateSpeed(e) {{
        speed = 1100 - parseInt(e.target.value); 
        if (isPlaying) {{
            pause();
            play();
        }}
    }}

    function updateLimit(e) {{
        topN = parseInt(e.target.value);
        updateFrame();
    }}

    function scrubTimeline(e) {{
        pause();
        currentIndex = parseInt(e.target.value);
        updateFrame();
    }}

    window.onload = init;
</script>

</body>
</html>
"""

    # Write HTML file
    output_html_path = "bar_chart_race.html"
    with open(output_html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    print(f"\n[SUCCESS] Interactive dashboard saved to: {output_html_path}")
    
    # 9. Open in browser automatically
    abs_path = os.path.abspath(output_html_path)
    print(f"[INFO] Automatically opening browser: {abs_path}")
    webbrowser.open("file://" + abs_path)

if __name__ == "__main__":
    main()
